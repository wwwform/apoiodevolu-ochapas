import streamlit as st
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
import time
import os
import io
from openpyxl.styles import numbers

# =====================================================
# CONFIGURAÇÃO STREAMLIT
# =====================================================
st.set_page_config(page_title="Sistema Chapas", layout="wide")

st.markdown("""
<style>
header[data-testid="stHeader"] {display: none;}
#MainMenu {visibility: hidden;}
footer {visibility: hidden;}
[data-testid="stToolbar"] {display: none;}
[data-testid="stDecoration"] {display: none;}
.stDeployButton {display:none;}
</style>
""", unsafe_allow_html=True)

# =====================================================
# GOOGLE SHEETS
# =====================================================
@st.cache_resource
def conectar_google():
    scope = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        st.secrets["gcp_service_account"], scope
    )
    client = gspread.authorize(creds)
    return client.open("BD_Fabrica_Geral")

def garantir_cabecalhos():
    sh = conectar_google()

    try:
        ws = sh.worksheet("Chapas_Producao")
    except:
        ws = sh.add_worksheet("Chapas_Producao", 1000, 20)

    if not ws.row_values(1):
        ws.append_row([
            "id","data_hora","lote","reserva","status_reserva",
            "cod_sap","descricao","qtd","peso_real",
            "largura_real_mm","largura_corte_mm",
            "tamanho_real_mm","tamanho_corte_mm",
            "peso_teorico","sucata"
        ])

    try:
        ws_l = sh.worksheet("Chapas_Lotes")
    except:
        ws_l = sh.add_worksheet("Chapas_Lotes", 1000, 5)

    if not ws_l.row_values(1):
        ws_l.append_row(["cod_sap","ultimo_numero"])

garantir_cabecalhos()

# =====================================================
# FUNÇÕES
# =====================================================
def normalizar_numero_br(v):
    if pd.isna(v):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except:
        return 0.0

def regra_300(mm):
    try:
        return (int(float(mm)) // 300) * 300
    except:
        return 0

@st.cache_data
def carregar_base_sap():
    path = "base_sap.xlsx"
    if not os.path.exists(path):
        return None

    df = pd.read_excel(path)
    df.columns = df.columns.str.strip().str.upper()

    col_prod = next((c for c in df.columns if 'PRODUTO' in c), None)
    col_peso = next((c for c in df.columns if 'PESO' in c and 'METRO' in c), None)

    if not col_prod or not col_peso:
        return None

    df['PRODUTO'] = pd.to_numeric(df[col_prod], errors='coerce').fillna(0).astype(int)
    df['PESO_FATOR'] = df[col_peso].apply(normalizar_numero_br)

    return df

# =====================================================
# APP
# =====================================================
st.sidebar.title("🔐 Acesso")
perfil = st.sidebar.radio("Perfil:", [
    "Operador (Chão de Fábrica)",
    "Administrador (Escritório)"
])

df_sap = carregar_base_sap()

# =====================================================
# OPERADOR
# =====================================================
if perfil == "Operador (Chão de Fábrica)":
    st.title("🏭 Chapas – Bipagem")

    if 'wizard' not in st.session_state:
        st.session_state.wizard = {}

    cod = st.text_input("BIPAR CÓDIGO SAP")

    if cod and df_sap is not None:
        try:
            cod = int(cod)
            prod = df_sap[df_sap['PRODUTO'] == cod].iloc[0]

            qtd = st.number_input("Qtd", min_value=1, step=1)
            peso_real = st.number_input("Peso Real (kg)", format="%.2f")
            larg = st.number_input("Largura Real (mm)", step=1)
            comp = st.number_input("Comprimento Real (mm)", step=1)

            lc = regra_300(larg)
            tc = regra_300(comp)

            peso_teorico = prod['PESO_FATOR'] * (lc/1000) * (tc/1000) * qtd
            sucata = peso_real - peso_teorico

            st.info(f"Peso Teórico: {peso_teorico:.2f} kg")

            if st.button("SALVAR"):
                sh = conectar_google()
                ws = sh.worksheet("Chapas_Producao")
                ws_l = sh.worksheet("Chapas_Lotes")

                try:
                    cell = ws_l.find(str(cod))
                    prox = int(ws_l.cell(cell.row, 2).value) + 1
                    ws_l.update_cell(cell.row, 2, prox)
                except:
                    prox = 1
                    ws_l.append_row([cod, prox])

                lote = f"BRASA{prox:05d}"

                ws.append_row([
                    int(datetime.now().timestamp()*1000),
                    datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    lote, "", "Pendente",
                    cod, prod['DESCRIÇÃO DO PRODUTO'],
                    qtd, peso_real,
                    larg, lc, comp, tc,
                    round(peso_teorico, 2),
                    round(sucata, 2)
                ])

                st.success("Salvo com sucesso")
                time.sleep(1)
                st.rerun()

        except:
            st.error("Produto não encontrado")

# =====================================================
# ADMINISTRADOR
# =====================================================
elif perfil == "Administrador (Escritório)":
    st.title("💻 Administração")

    sh = conectar_google()
    ws = sh.worksheet("Chapas_Producao")
    df = pd.DataFrame(ws.get_all_records())

    if not df.empty:
        for c in ['peso_real','peso_teorico','sucata','qtd']:
            df[c] = df[c].apply(normalizar_numero_br)

        # ===================== EXPORTAÇÃO CORRETA =====================
        export = df[[
            'lote','reserva','cod_sap','descricao',
            'status_reserva','qtd','peso_teorico'
        ]].copy()

        export.rename(columns={
            'peso_teorico':'Peso Lançamento (kg)'
        }, inplace=True)

        export['Peso Lançamento (kg)'] = export['Peso Lançamento (kg)'].round(2)

        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            export.to_excel(writer, index=False, sheet_name="Relatorio")
            wsx = writer.sheets["Relatorio"]

            col = export.columns.get_loc("Peso Lançamento (kg)") + 1
            for r in range(2, wsx.max_row + 1):
                wsx.cell(row=r, column=col).number_format = '0.00'

        st.download_button(
            "⬇️ Baixar Excel",
            buffer.getvalue(),
            "Relatorio_Chapas.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
